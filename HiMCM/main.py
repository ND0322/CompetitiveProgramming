import openpyxl

path = "TOP500_202406.xlsx"


wb = openpyxl.load_workbook(path)

sheet = wb.active



sm = 0
cnt = 0
for i in range(2, 502):
    if((sheet.cell(row = i, column = 21).value) is None):
        continue
    cnt += 1
    sm += (sheet.cell(row = i, column = 15).value) * 1000 / (sheet.cell(row = i, column = 21).value)

sm /= 1000

Eavg = sm/cnt

print("Avg Full Capacity kW: ", Eavg)
print("Extrapolated Total Full Capacity kW", Eavg * 500)

print("Avg Full Capacity tWh: ", Eavg * 365.25 * 24 / 1e9)
print("Extrapolated Total Full Capacity tWh: ", Eavg * 500 * 365.25 * 24 / 1e9)

